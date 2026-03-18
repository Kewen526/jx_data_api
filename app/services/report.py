#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
江鑫数据报表生成服务
支持生成：日报、周报、月报、自定义报表
"""

import os
import uuid
import json
import traceback
import warnings
from datetime import datetime, timedelta
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.database import get_db_pool

warnings.filterwarnings('ignore')


# ==================== 辅助函数 ====================
def ensure_temp_dir():
    """确保临时目录存在"""
    if not os.path.exists(settings.TEMP_DIR):
        os.makedirs(settings.TEMP_DIR)


def generate_temp_filename(prefix: str, ext: str = "xlsx") -> str:
    """生成唯一的临时文件名"""
    ensure_temp_dir()
    unique_id = uuid.uuid4().hex[:8]
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return os.path.join(settings.TEMP_DIR, f"{prefix}_{timestamp}_{unique_id}.{ext}")


def clean_sheet_name(name, max_length=31):
    """清理 Sheet 名称，符合 Excel 规范"""
    if not name:
        return "Sheet"
    illegal_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in illegal_chars:
        name = name.replace(char, '')
    if len(name) > max_length:
        name = name[:max_length]
    return name or "Sheet"


def apply_border(ws, min_row, max_row, min_col, max_col):
    """应用边框样式"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def safe_get_val(data, key, default=0):
    """安全获取值，处理None的情况"""
    if not data:
        return default
    value = data.get(key)
    if value is None:
        return default
    return value


def calc_rate(numerator, denominator):
    """计算比率"""
    if denominator and denominator > 0:
        return round(numerator / denominator * 100, 1)
    return 0


def calc_avg_price(total, count):
    """计算均价"""
    if count and count > 0:
        return round(total / count, 2)
    return 0


# ==================== 数据查询辅助函数 ====================
def get_shop_info_mapping(accounts: Optional[List[str]] = None) -> dict:
    """获取门店信息映射"""
    db = get_db_pool()
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sql = """
        SELECT
            pa.account,
            pa.stores_json,
            pa.sales_name,
            pa.city_name,
            pa.operator_id,
            su.name as operator_name
        FROM platform_accounts pa
        LEFT JOIN saas_users su ON pa.operator_id = su.id
        WHERE pa.stores_json IS NOT NULL
        """
        params = []
        if accounts:
            placeholders = ','.join(['%s'] * len(accounts))
            sql += f" AND pa.account IN ({placeholders})"
            params = accounts

        cursor.execute(sql, params)
        account_results = cursor.fetchall()

        shop_mapping = {}
        for account in account_results:
            stores_json = account.get('stores_json')
            sales_name = account.get('sales_name', '')
            city_name = account.get('city_name', '')
            operator_name = account.get('operator_name', '')

            if stores_json:
                try:
                    if isinstance(stores_json, str):
                        stores = json.loads(stores_json)
                    else:
                        stores = stores_json

                    if isinstance(stores, list):
                        for store in stores:
                            if isinstance(store, dict):
                                shop_id = str(store.get('shop_id', ''))
                                if shop_id:
                                    shop_mapping[shop_id] = {
                                        'operator': operator_name or '',
                                        'sales': sales_name or '',
                                        'city': city_name or ''
                                    }
                except (json.JSONDecodeError, TypeError):
                    pass

        return shop_mapping
    finally:
        cursor.close()
        conn.close()


def get_region_info_mapping(accounts: Optional[List[str]] = None) -> dict:
    """获取商圈信息映射"""
    db = get_db_pool()
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sql = """
        SELECT pa.compareRegions_json
        FROM platform_accounts pa
        WHERE pa.compareRegions_json IS NOT NULL
        """
        params = []
        if accounts:
            placeholders = ','.join(['%s'] * len(accounts))
            sql += f" AND pa.account IN ({placeholders})"
            params = accounts

        cursor.execute(sql, params)
        account_results = cursor.fetchall()

        region_mapping = {}
        for account in account_results:
            regions_json = account.get('compareRegions_json')
            if regions_json:
                try:
                    if isinstance(regions_json, str):
                        regions = json.loads(regions_json)
                    else:
                        regions = regions_json

                    if isinstance(regions, dict):
                        for shop_id, shop_data in regions.items():
                            if isinstance(shop_data, dict):
                                regions_data = shop_data.get('regions', {})
                                if isinstance(regions_data, dict):
                                    city_info = regions_data.get('city', {})
                                    district_info = regions_data.get('district', {})
                                    business_info = regions_data.get('business', {})

                                    region_mapping[str(shop_id)] = {
                                        'city': city_info.get('regionName', '') if isinstance(city_info, dict) else '',
                                        'district': district_info.get('regionName', '') if isinstance(district_info, dict) else '',
                                        'business': business_info.get('regionName', '') if isinstance(business_info, dict) else ''
                                    }
                except (json.JSONDecodeError, TypeError):
                    pass

        return region_mapping
    finally:
        cursor.close()
        conn.close()


def get_coupon_orders_last_7days(shop_id, report_date):
    """获取近7天优惠码订单总数"""
    db = get_db_pool()
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        end_date = datetime.strptime(report_date, '%Y-%m-%d')
        start_date = end_date - timedelta(days=6)

        sql = """
        SELECT COALESCE(SUM(coupon_pay_order_count), 0) as total
        FROM kewen_daily_report
        WHERE shop_id = %s AND report_date BETWEEN %s AND %s
        """
        cursor.execute(sql, (shop_id, start_date.strftime('%Y-%m-%d'), report_date))
        result = cursor.fetchone()
        return int(result['total']) if result and result['total'] else 0
    finally:
        cursor.close()
        conn.close()


def get_ad_orders_today(shop_id, report_date):
    """获取当天广告单数量"""
    db = get_db_pool()
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sql = """
        SELECT COALESCE(ad_order_count, 0) as total
        FROM store_stats
        WHERE store_id = %s AND date = %s
        """
        cursor.execute(sql, (shop_id, report_date))
        result = cursor.fetchone()
        return int(result['total']) if result and result['total'] else 0
    finally:
        cursor.close()
        conn.close()


# ==================== 核心功能：生成日报 ====================
def generate_daily_report(report_date: str, accounts: Optional[List[str]] = None) -> str:
    """
    生成日报
    参数:
        report_date: 报表日期，格式: 'YYYY-MM-DD'
        accounts: 门店账号列表，如["13718175572a","19318574226a"]
    返回:
        生成的文件路径
    """
    db = get_db_pool()

    # 如果指定了accounts，先获取对应的shop_id列表
    shop_ids_filter = None
    if accounts:
        conn_temp = db.get_connection()
        cursor_temp = conn_temp.cursor(dictionary=True)
        try:
            placeholders = ','.join(['%s'] * len(accounts))
            sql_accounts = f"""
            SELECT stores_json FROM platform_accounts WHERE account IN ({placeholders})
            """
            cursor_temp.execute(sql_accounts, accounts)
            account_data = cursor_temp.fetchall()

            shop_ids_filter = []
            shop_name_fallback = {}
            for acc in account_data:
                stores_json = acc.get('stores_json')
                if stores_json:
                    try:
                        if isinstance(stores_json, str):
                            stores = json.loads(stores_json)
                        else:
                            stores = stores_json

                        if isinstance(stores, list):
                            for store in stores:
                                if isinstance(store, dict):
                                    shop_id = str(store.get('shop_id', ''))
                                    if shop_id:
                                        shop_ids_filter.append(shop_id)
                                        shop_name_fallback[shop_id] = (
                                            store.get('shop_name') or store.get('name') or f'门店{shop_id}'
                                        )
                    except (json.JSONDecodeError, TypeError):
                        pass
        finally:
            cursor_temp.close()
            conn_temp.close()

    # 获取门店信息映射
    shop_mapping = get_shop_info_mapping(accounts)
    region_mapping = get_region_info_mapping(accounts)

    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if shop_ids_filter is not None and not shop_ids_filter:
            # 账号存在但无对应门店，直接返回空结果
            rows = []
        elif shop_ids_filter:
            # 以传入的店铺ID列表为驱动表，LEFT JOIN三张数据表
            # 无论哪张表有无数据，所有传入的店铺都会出现在结果中
            union_parts = []
            union_params = []
            for sid in shop_ids_filter:
                union_parts.append("SELECT %s AS shop_id, %s AS shop_name_hint")
                union_params.extend([sid, shop_name_fallback.get(sid, f'门店{sid}')])
            shop_subquery = ' UNION ALL '.join(union_parts)

            sql = f"""
            SELECT
                sl.shop_id,
                COALESCE(k.shop_name, p.shop_name, sl.shop_name_hint) AS shop_name,
                k.exposure_users, k.visit_users, k.order_users,
                k.verify_person_count AS verify_users,
                k.order_coupon_count, k.verify_coupon_count,
                k.promotion_cost, k.new_good_review_count,
                sds.new_review AS new_review_count,
                sds.new_favorite AS new_collect_users,
                sds.online_consult AS consult_users,
                k.intent_rate,
                k.order_sale_amount, k.verify_sale_amount, k.verify_after_discount,
                sds.phone AS phone_clicks,
                sds.address AS address_clicks,
                p.click_avg_price, p.order_count AS promotion_order_count,
                s.order_user_rank, s.verify_amount_rank,
                sds.new_checkin AS checkin_count, s.ad_balance, s.ad_order_count, s.is_force_offline
            FROM ({shop_subquery}) AS sl
            LEFT JOIN kewen_daily_report k ON sl.shop_id = k.shop_id AND k.report_date = %s
            LEFT JOIN promotion_daily_report p ON sl.shop_id = p.shop_id AND p.report_date = %s
            LEFT JOIN store_stats s ON sl.shop_id = s.store_id AND s.date = %s
            LEFT JOIN store_daily_stats sds ON sl.shop_id = sds.store_id AND sds.data_date = %s
            ORDER BY sl.shop_id
            """
            params = union_params + [report_date, report_date, report_date, report_date]
            cursor.execute(sql, params)
            rows = list(cursor.fetchall())
            for row in rows:
                row['report_date'] = report_date
        else:
            # 未指定 accounts，查询该日期下所有门店数据（保持原有行为）
            sql = """
            SELECT
                k.report_date, k.shop_id, k.shop_name,
                k.exposure_users, k.visit_users, k.order_users,
                k.verify_person_count AS verify_users,
                k.order_coupon_count, k.verify_coupon_count,
                k.promotion_cost, k.new_good_review_count,
                sds.new_review AS new_review_count,
                sds.new_favorite AS new_collect_users,
                sds.online_consult AS consult_users,
                k.intent_rate,
                k.order_sale_amount, k.verify_sale_amount, k.verify_after_discount,
                sds.phone AS phone_clicks,
                sds.address AS address_clicks,
                p.click_avg_price, p.order_count AS promotion_order_count,
                s.order_user_rank, s.verify_amount_rank,
                sds.new_checkin AS checkin_count, s.ad_balance, s.ad_order_count, s.is_force_offline
            FROM kewen_daily_report k
            LEFT JOIN promotion_daily_report p ON k.shop_id = p.shop_id AND k.report_date = p.report_date
            LEFT JOIN store_stats s ON k.shop_id = s.store_id AND k.report_date = s.date
            LEFT JOIN store_daily_stats sds ON k.shop_id = sds.store_id AND k.report_date = sds.data_date
            WHERE k.report_date = %s
            ORDER BY k.shop_id
            """
            cursor.execute(sql, [report_date])
            rows = list(cursor.fetchall())

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "汇总"

        # 格式化日期
        date_obj = datetime.strptime(report_date, '%Y-%m-%d')
        weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        weekday = weekday_names[date_obj.weekday()]
        date_str = date_obj.strftime('%m月%d日')
        date_short = date_obj.strftime('%m/%d')

        # 汇总表头
        summary_headers = [
            '星期', '日期', '序号', '运营', '城市', '销售', '门店',
            '曝光人数', '访问人数', '下单人数', '核销人数', '下单券数', '核销券数',
            '电话点击', '地址点击', '推广通消耗', '好评', '意向转化率',
            '下单售价金额', '核销售价金额', '优惠后核销金额',
            '下单人数商圈排名', '核销金额商圈排名'
        ]
        ws_summary.append(summary_headers)

        # 汇总表头样式
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=10)
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet_names_used = {}

        for idx, row in enumerate(rows, start=1):
            shop_id = str(row['shop_id'])
            shop_name = row['shop_name'] or f'门店{shop_id}'

            shop_info = shop_mapping.get(shop_id, {})
            operator = shop_info.get('operator', '')
            sales = shop_info.get('sales', '')
            city = shop_info.get('city', '')

            region_info = region_mapping.get(shop_id, {})
            region_city = region_info.get('city', city)
            region_district = region_info.get('district', '')
            region_business = region_info.get('business', '')

            order_rank = row['order_user_rank']
            verify_rank = row['verify_amount_rank']
            order_rank_str = f"第{order_rank}名" if order_rank and order_rank < 100 else ("大于100名" if order_rank and order_rank >= 100 else "--")
            verify_rank_str = f"第{verify_rank}名" if verify_rank and verify_rank < 100 else ("大于100名" if verify_rank and verify_rank >= 100 else "--")

            summary_row = [
                weekday, date_str, idx, operator, city, sales, shop_name,
                row['exposure_users'] or 0, row['visit_users'] or 0,
                row['order_users'] or 0, row['verify_users'] or 0,
                row['order_coupon_count'] or 0, row['verify_coupon_count'] or 0,
                row['phone_clicks'] or 0, row['address_clicks'] or 0,
                round(row['promotion_cost'], 2) if row['promotion_cost'] else 0,
                row['new_good_review_count'] or 0, row['intent_rate'] or '0%',
                round(row['order_sale_amount'], 2) if row['order_sale_amount'] else 0,
                round(row['verify_sale_amount'], 2) if row['verify_sale_amount'] else 0,
                round(row['verify_after_discount'], 2) if row['verify_after_discount'] else 0,
                order_rank_str, verify_rank_str
            ]
            ws_summary.append(summary_row)

            # 创建门店详细Sheet
            sheet_name = clean_sheet_name(shop_name)
            if sheet_name in sheet_names_used:
                sheet_names_used[sheet_name] += 1
                sheet_name = f"{sheet_name[:28]}_{sheet_names_used[sheet_name]}"
            else:
                sheet_names_used[sheet_name] = 1

            ws_detail = wb.create_sheet(title=sheet_name)

            order_users = row['order_users'] or 0
            verify_users = row['verify_users'] or 0
            new_review_count = row['new_review_count'] or 0
            new_collect_users = row['new_collect_users'] or 0

            review_rate = (new_review_count / verify_users * 100) if verify_users > 0 else 0
            review_rate_str = f"{review_rate:.1f}%"
            review_qualified = "达标" if review_rate >= 30 else "未达标"

            collect_rate = (new_collect_users / order_users * 100) if order_users > 0 else 0
            collect_rate_str = f"{collect_rate:.1f}%"
            collect_qualified = "达标" if collect_rate >= 40 else "未达标"

            coupon_7days = get_coupon_orders_last_7days(shop_id, report_date)
            coupon_qualified = "达标" if coupon_7days >= 10 else "未达标"

            ad_today = get_ad_orders_today(shop_id, report_date)
            ad_qualified = "达标" if ad_today >= 1 else "未达标"

            is_force_offline = row['is_force_offline'] or 0
            status_info = f"警告：有{is_force_offline}个团单被强制下线！" if is_force_offline > 0 else "今天邮件已查看，无违规无异常。"

            region_display = f"{region_city} | {region_district} | {region_business}" if region_business else city
            order_rank_display = f"{region_display}：第{order_rank}名" if order_rank and order_rank < 100 else f"{region_display}：大于100名"
            verify_rank_display = f"{region_display}：第{verify_rank}名" if verify_rank and verify_rank < 100 else f"{region_display}：大于100名"

            detail_data = [
                [shop_name, status_info, ''],
                [f"数据报表", f"日期({date_short})", ''],
                ['【美团点评广告结果数据】', '', ''],
                ['曝光人数：', row['exposure_users'] or 0, ''],
                ['访问人数：', row['visit_users'] or 0, ''],
                ['下单人数：', row['order_users'] or 0, ''],
                ['下单券数：', row['order_coupon_count'] or 0, ''],
                ['核销人数：', row['verify_users'] or 0, ''],
                ['核销券数：', row['verify_coupon_count'] or 0, ''],
                ['电话点击：', row['phone_clicks'] or 0, ''],
                ['地址点击：', row['address_clicks'] or 0, ''],
                ['在线咨询：', row['consult_users'] or 0, ''],
                ['', '', ''],
                ['【店内干预数据】', '', ''],
                ['新增收藏：', row['new_collect_users'] or 0, ''],
                ['新增打卡：', row['checkin_count'] or 0, ''],
                ['新增评价：', row['new_review_count'] or 0, ''],
                ['', '', ''],
                ['【推广通数据】', '', ''],
                ['推广通消耗：', round(row['promotion_cost'], 2) if row['promotion_cost'] else 0, ''],
                ['推广通点击单价：', round(row['click_avg_price'], 2) if row['click_avg_price'] else 0, ''],
                ['推广通下单量：', row['promotion_order_count'] or 0, ''],
                ['推广通余额：', round(row['ad_balance'], 2) if row['ad_balance'] else 0, ''],
                ['', '', ''],
                [f'留评率（30%达标）：', review_rate_str, review_qualified],
                [f'收藏率（40%达标）：', collect_rate_str, collect_qualified],
                [f'近7天优惠码订单是否达标：', coupon_7days, coupon_qualified],
                [f'广告单：', f"当天{ad_today}单", ad_qualified],
                ['', '', ''],
                ['下单售价金额：', round(row['order_sale_amount'], 2) if row['order_sale_amount'] else 0, ''],
                ['核销售价金额：', round(row['verify_sale_amount'], 2) if row['verify_sale_amount'] else 0, ''],
                ['下单人数商圈排名：', order_rank_display, ''],
                ['核销金额商圈排名：', verify_rank_display, ''],
                ['', '', ''],
                ['团单被强制下线数量：', is_force_offline, ''],
                ['', '', ''],
                ['运营：', operator, ''],
                ['销售：', sales, ''],
                ['城市：', city, ''],
            ]

            for row_data in detail_data:
                ws_detail.append(row_data)

            ws_detail.column_dimensions['A'].width = 40
            ws_detail.column_dimensions['B'].width = 30
            ws_detail.column_dimensions['C'].width = 15

            for row_num in range(1, len(detail_data) + 1):
                for col_num in range(1, 4):
                    ws_detail.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

            ws_detail['A1'].font = Font(bold=True, size=12)
            ws_detail['B1'].font = Font(bold=True, size=10, color="FF0000" if is_force_offline > 0 else "008000")

            section_rows = [3, 14, 19]
            for r in section_rows:
                ws_detail.cell(row=r, column=1).font = Font(bold=True, size=10, color="0066CC")

            qualified_rows = [25, 26, 27, 28]
            for r in qualified_rows:
                cell = ws_detail.cell(row=r, column=3)
                if cell.value == "未达标":
                    cell.font = Font(bold=True, color="FF0000")
                elif cell.value == "达标":
                    cell.font = Font(bold=True, color="008000")

            apply_border(ws_detail, 1, len(detail_data), 1, 3)

        summary_widths = [6, 8, 5, 12, 8, 8, 46, 10, 10, 10, 10, 10, 10, 10, 10, 12, 8, 12, 12, 12, 12, 14, 14]
        for col_idx, width in enumerate(summary_widths, start=1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

        for row in ws_summary.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(summary_headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        apply_border(ws_summary, 1, len(rows) + 1, 1, len(summary_headers))

        output_filename = generate_temp_filename(f"日报_{report_date.replace('-', '')}")
        wb.save(output_filename)

        return output_filename

    finally:
        cursor.close()
        conn.close()


# ==================== 辅助：按时间段查询三表数据并合并 ====================
def _fetch_period_data(cursor, start_date, end_date, shop_ids_filter, shop_name_fallback):
    """
    分别查询 kewen_daily_report、promotion_daily_report、store_stats，
    按店铺ID合并后返回。

    当 shop_ids_filter 不为 None 时：以传入的店铺列表为准，
    确保所有店铺都出现在结果中（某表无数据则对应字段为 None）。
    当 shop_ids_filter 为 None 时：以 kewen_daily_report 的店铺为准（原有行为）。
    """
    if shop_ids_filter is not None and not shop_ids_filter:
        return {}

    def shop_clause(id_col, ids):
        if not ids:
            return "", []
        ph = ','.join(['%s'] * len(ids))
        return f" AND {id_col} IN ({ph})", list(ids)

    k_clause, k_params = shop_clause('shop_id', shop_ids_filter)
    p_clause, p_params = shop_clause('shop_id', shop_ids_filter)
    s_clause, s_params = shop_clause('store_id', shop_ids_filter)

    # kewen_daily_report
    sql_k = f"""
    SELECT shop_id, MAX(shop_name) AS shop_name,
        SUM(verify_after_discount) AS verify_after_discount,
        SUM(exposure_users) AS exposure_users,
        SUM(visit_users) AS visit_users,
        SUM(order_users) AS order_users,
        SUM(order_coupon_count) AS order_coupon_count,
        SUM(verify_person_count) AS verify_users,
        SUM(verify_coupon_count) AS verify_coupon_count,
        SUM(order_sale_amount) AS order_sale_amount,
        SUM(verify_sale_amount) AS verify_sale_amount,
        SUM(coupon_pay_order_count) AS coupon_orders,
        SUM(promotion_cost) AS promotion_cost,
        SUM(promotion_exposure_count) AS promotion_exposure,
        SUM(promotion_click_count) AS promotion_clicks,
        SUM(consult_users) AS consult_users,
        SUM(new_collect_users) AS new_collect,
        SUM(new_good_review_count) AS new_good_reviews,
        SUM(new_review_count) AS new_reviews
    FROM kewen_daily_report
    WHERE report_date BETWEEN %s AND %s{k_clause}
    GROUP BY shop_id
    """
    cursor.execute(sql_k, [start_date, end_date] + k_params)
    kewen_data = {str(r['shop_id']): r for r in cursor.fetchall()}

    # promotion_daily_report
    sql_p = f"""
    SELECT shop_id, MAX(shop_name) AS shop_name,
        SUM(view_phone_count) AS phone_clicks,
        SUM(order_count) AS promotion_orders,
        SUM(view_groupbuy_count) AS view_groupbuy,
        SUM(view_phone_count) AS view_phone,
        SUM(view_address_count) AS address_clicks
    FROM promotion_daily_report
    WHERE report_date BETWEEN %s AND %s{p_clause}
    GROUP BY shop_id
    """
    cursor.execute(sql_p, [start_date, end_date] + p_params)
    promo_data = {str(r['shop_id']): r for r in cursor.fetchall()}

    # store_stats
    sql_s = f"""
    SELECT store_id,
        SUM(checkin_count) AS checkin_count
    FROM store_stats
    WHERE date BETWEEN %s AND %s{s_clause}
    GROUP BY store_id
    """
    cursor.execute(sql_s, [start_date, end_date] + s_params)
    stats_data = {str(r['store_id']): r for r in cursor.fetchall()}

    # store_daily_stats（电话点击、地址点击、在线咨询、新增收藏、新增打卡、新增评价）
    sds_clause, sds_params = shop_clause('store_id', shop_ids_filter)
    sql_sds = f"""
    SELECT store_id,
        SUM(phone) AS phone_clicks,
        SUM(address) AS address_clicks,
        SUM(online_consult) AS consult_users,
        SUM(new_favorite) AS new_collect,
        SUM(new_checkin) AS checkin_count,
        SUM(new_review) AS new_reviews
    FROM store_daily_stats
    WHERE data_date BETWEEN %s AND %s{sds_clause}
    GROUP BY store_id
    """
    cursor.execute(sql_sds, [start_date, end_date] + sds_params)
    sds_data = {str(r['store_id']): r for r in cursor.fetchall()}

    # 确定需要输出的店铺集合
    if shop_ids_filter is not None:
        # 以传入列表为准，确保所有店铺都出现
        all_sids = {str(s) for s in shop_ids_filter}
    else:
        # 未指定过滤条件，以 kewen 数据为准（保持原有行为）
        all_sids = set(kewen_data.keys())

    merged = {}
    for sid in all_sids:
        k = kewen_data.get(sid, {})
        p = promo_data.get(sid, {})
        s = stats_data.get(sid, {})
        sds = sds_data.get(sid, {})
        shop_name = k.get('shop_name') or p.get('shop_name') or shop_name_fallback.get(sid, f'门店{sid}')
        merged[sid] = {
            'shop_id': sid,
            'shop_name': shop_name,
            'verify_after_discount': k.get('verify_after_discount'),
            'exposure_users': k.get('exposure_users'),
            'visit_users': k.get('visit_users'),
            'order_users': k.get('order_users'),
            'order_coupon_count': k.get('order_coupon_count'),
            'verify_users': k.get('verify_users'),
            'verify_coupon_count': k.get('verify_coupon_count'),
            'order_sale_amount': k.get('order_sale_amount'),
            'verify_sale_amount': k.get('verify_sale_amount'),
            'coupon_orders': k.get('coupon_orders'),
            'promotion_cost': k.get('promotion_cost'),
            'promotion_exposure': k.get('promotion_exposure'),
            'promotion_clicks': k.get('promotion_clicks'),
            'consult_users': sds.get('consult_users'),
            'new_collect': sds.get('new_collect'),
            'new_good_reviews': k.get('new_good_reviews'),
            'new_reviews': sds.get('new_reviews'),
            'phone_clicks': sds.get('phone_clicks'),
            'promotion_orders': p.get('promotion_orders'),
            'view_groupbuy': p.get('view_groupbuy'),
            'view_phone': p.get('view_phone'),
            'address_clicks': sds.get('address_clicks'),
            'checkin_count': sds.get('checkin_count'),
        }

    return merged


# ==================== 核心功能：生成周报 ====================
def generate_weekly_report(
    week1_start: str,
    week1_end: str,
    week2_start: str,
    week2_end: str,
    accounts: Optional[List[str]] = None,
    shop_id: Optional[str] = None
) -> str:
    """
    生成周报（两周对比）
    """
    db = get_db_pool()

    # 如果指定了shop_id，直接使用该shop_id过滤
    # 否则如果指定了accounts，获取对应的shop_id列表
    shop_ids_filter = None
    shop_name_fallback = {}
    if shop_id:
        # 直接使用传入的单个shop_id
        shop_ids_filter = [shop_id]
    elif accounts:
        conn_temp = db.get_connection()
        cursor_temp = conn_temp.cursor(dictionary=True)
        try:
            placeholders = ','.join(['%s'] * len(accounts))
            sql_accounts = f"""
            SELECT stores_json FROM platform_accounts WHERE account IN ({placeholders})
            """
            cursor_temp.execute(sql_accounts, accounts)
            account_data = cursor_temp.fetchall()

            shop_ids_filter = []
            for acc in account_data:
                stores_json = acc.get('stores_json')
                if stores_json:
                    try:
                        if isinstance(stores_json, str):
                            stores = json.loads(stores_json)
                        else:
                            stores = stores_json

                        if isinstance(stores, list):
                            for store in stores:
                                if isinstance(store, dict):
                                    sid = str(store.get('shop_id', ''))
                                    if sid:
                                        shop_ids_filter.append(sid)
                                        shop_name_fallback[sid] = (
                                            store.get('shop_name') or store.get('name') or f'门店{sid}'
                                        )
                    except (json.JSONDecodeError, TypeError):
                        pass
        finally:
            cursor_temp.close()
            conn_temp.close()

    shop_mapping = get_shop_info_mapping(accounts)

    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 分别查询三张表并合并，确保所有传入的店铺都出现在结果中
        week1_data = _fetch_period_data(cursor, week1_start, week1_end, shop_ids_filter, shop_name_fallback)
        week2_data = _fetch_period_data(cursor, week2_start, week2_end, shop_ids_filter, shop_name_fallback)

        all_shop_ids = set(week1_data.keys()) | set(week2_data.keys())

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "汇总"

        week1_str = f"{datetime.strptime(week1_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(week1_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"
        week2_str = f"{datetime.strptime(week2_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(week2_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"

        sheet_names_used = {}
        seq_num = 1

        for shop_id in sorted(all_shop_ids):
            w1 = week1_data.get(shop_id, {})
            w2 = week2_data.get(shop_id, {})
            shop_name = w2.get('shop_name') or w1.get('shop_name') or shop_name_fallback.get(str(shop_id), f'门店{shop_id}')

            shop_id_str = str(shop_id)
            shop_info = shop_mapping.get(shop_id_str, {})
            operator = shop_info.get('operator', '--') or '--'
            sales = shop_info.get('sales', '--') or '--'
            city = shop_info.get('city', '--') or '--'

            # 周1指标
            w1_verify_discount = safe_get_val(w1, 'verify_after_discount')
            w1_exposure = safe_get_val(w1, 'exposure_users')
            w1_visit = safe_get_val(w1, 'visit_users')
            w1_order_users = safe_get_val(w1, 'order_users')
            w1_order_coupons = safe_get_val(w1, 'order_coupon_count')
            w1_verify_users = safe_get_val(w1, 'verify_users')
            w1_verify_coupons = safe_get_val(w1, 'verify_coupon_count')
            w1_order_amount = safe_get_val(w1, 'order_sale_amount')
            w1_verify_amount = safe_get_val(w1, 'verify_sale_amount')
            w1_coupon_orders = safe_get_val(w1, 'coupon_orders')
            w1_phone_clicks = safe_get_val(w1, 'phone_clicks')

            w1_exposure_rate = f"{calc_rate(w1_visit, w1_exposure)}%"
            w1_order_rate = f"{calc_rate(w1_order_users, w1_visit)}%"
            w1_avg_price = calc_avg_price(w1_verify_discount, w1_verify_users)

            # 周2指标
            w2_verify_discount = safe_get_val(w2, 'verify_after_discount')
            w2_exposure = safe_get_val(w2, 'exposure_users')
            w2_visit = safe_get_val(w2, 'visit_users')
            w2_order_users = safe_get_val(w2, 'order_users')
            w2_order_coupons = safe_get_val(w2, 'order_coupon_count')
            w2_verify_users = safe_get_val(w2, 'verify_users')
            w2_verify_coupons = safe_get_val(w2, 'verify_coupon_count')
            w2_order_amount = safe_get_val(w2, 'order_sale_amount')
            w2_verify_amount = safe_get_val(w2, 'verify_sale_amount')
            w2_coupon_orders = safe_get_val(w2, 'coupon_orders')
            w2_phone_clicks = safe_get_val(w2, 'phone_clicks')

            w2_exposure_rate = f"{calc_rate(w2_visit, w2_exposure)}%"
            w2_order_rate = f"{calc_rate(w2_order_users, w2_visit)}%"
            w2_avg_price = calc_avg_price(w2_verify_discount, w2_verify_users)

            # 差值
            diff_verify_discount = round(w2_verify_discount - w1_verify_discount, 2)
            diff_exposure = w2_exposure - w1_exposure
            diff_visit = w2_visit - w1_visit

            def calc_rate_diff(rate1_str, rate2_str):
                val1 = float(rate1_str.rstrip('%')) if rate1_str != '0%' else 0
                val2 = float(rate2_str.rstrip('%')) if rate2_str != '0%' else 0
                return f"{round(val2 - val1, 1)}%"

            diff_exposure_rate = calc_rate_diff(w1_exposure_rate, w2_exposure_rate)
            diff_order_users = w2_order_users - w1_order_users
            diff_order_coupons = w2_order_coupons - w1_order_coupons
            diff_order_rate = calc_rate_diff(w1_order_rate, w2_order_rate)
            diff_verify_users = w2_verify_users - w1_verify_users
            diff_verify_coupons = w2_verify_coupons - w1_verify_coupons
            diff_order_amount = round(w2_order_amount - w1_order_amount, 2)
            diff_verify_amount = round(w2_verify_amount - w1_verify_amount, 2)
            diff_coupon_orders = w2_coupon_orders - w1_coupon_orders
            diff_phone_clicks = w2_phone_clicks - w1_phone_clicks
            diff_avg_price = round(w2_avg_price - w1_avg_price, 2)

            # 推广通数据
            w1_promo_cost = safe_get_val(w1, 'promotion_cost')
            w1_promo_exposure = safe_get_val(w1, 'promotion_exposure')
            w1_promo_clicks = safe_get_val(w1, 'promotion_clicks')
            w1_promo_orders = safe_get_val(w1, 'promotion_orders')
            w1_view_groupbuy = safe_get_val(w1, 'view_groupbuy')
            w1_view_phone = safe_get_val(w1, 'view_phone')
            w1_consult = safe_get_val(w1, 'consult_users')
            w1_address = safe_get_val(w1, 'address_clicks')
            w1_collect = safe_get_val(w1, 'new_collect')
            w1_good_reviews = safe_get_val(w1, 'new_good_reviews')
            w1_click_price = calc_avg_price(w1_promo_cost, w1_promo_clicks)
            w1_promo_rate = f"{calc_rate(w1_promo_orders, w1_promo_clicks)}%"
            w1_collect_rate = f"{calc_rate(w1_collect, w1_visit)}%"
            w1_review_rate = f"{calc_rate(w1_good_reviews, w1_verify_users)}%"

            w2_promo_cost = safe_get_val(w2, 'promotion_cost')
            w2_promo_exposure = safe_get_val(w2, 'promotion_exposure')
            w2_promo_clicks = safe_get_val(w2, 'promotion_clicks')
            w2_promo_orders = safe_get_val(w2, 'promotion_orders')
            w2_view_groupbuy = safe_get_val(w2, 'view_groupbuy')
            w2_view_phone = safe_get_val(w2, 'view_phone')
            w2_consult = safe_get_val(w2, 'consult_users')
            w2_address = safe_get_val(w2, 'address_clicks')
            w2_collect = safe_get_val(w2, 'new_collect')
            w2_good_reviews = safe_get_val(w2, 'new_good_reviews')
            w2_click_price = calc_avg_price(w2_promo_cost, w2_promo_clicks)
            w2_promo_rate = f"{calc_rate(w2_promo_orders, w2_promo_clicks)}%"
            w2_collect_rate = f"{calc_rate(w2_collect, w2_visit)}%"
            w2_review_rate = f"{calc_rate(w2_good_reviews, w2_verify_users)}%"

            diff_promo_cost = round(w2_promo_cost - w1_promo_cost, 2)
            diff_promo_exposure = w2_promo_exposure - w1_promo_exposure
            diff_promo_clicks = w2_promo_clicks - w1_promo_clicks
            diff_click_price = round(w2_click_price - w1_click_price, 2)
            diff_promo_orders = w2_promo_orders - w1_promo_orders
            diff_promo_rate = calc_rate_diff(w1_promo_rate, w2_promo_rate)
            diff_view_groupbuy = w2_view_groupbuy - w1_view_groupbuy
            diff_view_phone = w2_view_phone - w1_view_phone
            diff_consult = w2_consult - w1_consult
            diff_address = w2_address - w1_address
            diff_collect = w2_collect - w1_collect
            diff_collect_rate = calc_rate_diff(w1_collect_rate, w2_collect_rate)
            diff_good_reviews = w2_good_reviews - w1_good_reviews
            diff_review_rate = calc_rate_diff(w1_review_rate, w2_review_rate)

            # 汇总表8行结构
            header_row1 = ['序号', '运营', '城市', '销售', '门店', '数据周期', '优惠后核销额', '曝光人数', '访问人数', '曝光访问转化率', '下单人数', '下单券数', '下单转化率', '核销人数', '核销券数', '下单售价金额', '核销售价金额', '优惠码订单', '电话点击', '客单价']
            ws_summary.append(header_row1)

            row2 = [seq_num, operator, city, sales, shop_name, week1_str, round(w1_verify_discount, 2), w1_exposure, w1_visit, w1_exposure_rate, w1_order_users, w1_order_coupons, w1_order_rate, w1_verify_users, w1_verify_coupons, round(w1_order_amount, 2), round(w1_verify_amount, 2), w1_coupon_orders, w1_phone_clicks, w1_avg_price]
            ws_summary.append(row2)

            row3 = ['', '', '', '', '', week2_str, round(w2_verify_discount, 2), w2_exposure, w2_visit, w2_exposure_rate, w2_order_users, w2_order_coupons, w2_order_rate, w2_verify_users, w2_verify_coupons, round(w2_order_amount, 2), round(w2_verify_amount, 2), w2_coupon_orders, w2_phone_clicks, w2_avg_price]
            ws_summary.append(row3)

            row4 = ['', '', '', '', '', '差值', diff_verify_discount, diff_exposure, diff_visit, diff_exposure_rate, diff_order_users, diff_order_coupons, diff_order_rate, diff_verify_users, diff_verify_coupons, diff_order_amount, diff_verify_amount, diff_coupon_orders, diff_phone_clicks, diff_avg_price]
            ws_summary.append(row4)

            header_row2 = ['', '', '', '', '', '数据周期', '推广通花费', '推广通曝光', '推广通点击', '推广通点击均价', '推广通订单量', '推广通下单转化率', '推广通查看团购', '推广通查看电话', '在线咨询', '地址点击', '门店收藏', '收藏率', '新增好评数', '留评率']
            ws_summary.append(header_row2)

            row6 = ['', '', '', '', '', week1_str, round(w1_promo_cost, 2), w1_promo_exposure, w1_promo_clicks, w1_click_price, w1_promo_orders, w1_promo_rate, w1_view_groupbuy, w1_view_phone, w1_consult, w1_address, w1_collect, w1_collect_rate, w1_good_reviews, w1_review_rate]
            ws_summary.append(row6)

            row7 = ['', '', '', '', '', week2_str, round(w2_promo_cost, 2), w2_promo_exposure, w2_promo_clicks, w2_click_price, w2_promo_orders, w2_promo_rate, w2_view_groupbuy, w2_view_phone, w2_consult, w2_address, w2_collect, w2_collect_rate, w2_good_reviews, w2_review_rate]
            ws_summary.append(row7)

            row8 = ['', '', '', '', '', '差值', diff_promo_cost, diff_promo_exposure, diff_promo_clicks, diff_click_price, diff_promo_orders, diff_promo_rate, diff_view_groupbuy, diff_view_phone, diff_consult, diff_address, diff_collect, diff_collect_rate, diff_good_reviews, diff_review_rate]
            ws_summary.append(row8)

            seq_num += 1

        # 设置样式
        ws_summary.column_dimensions['A'].width = 8
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 10
        ws_summary.column_dimensions['E'].width = 78
        ws_summary.column_dimensions['F'].width = 26
        for i in range(7, 21):
            ws_summary.column_dimensions[get_column_letter(i)].width = 15

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        red_font = Font(color="FF0000")

        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, min_col=1, max_col=20):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column == 1 and cell.value == '序号':
                    for c in row:
                        c.font = Font(bold=True, size=10)
                        c.fill = green_fill
                elif cell.column == 6 and cell.value == '数据周期' and ws_summary.cell(row[0].row, 1).value == '':
                    for c in row:
                        c.font = Font(bold=True, size=10)
                        c.fill = green_fill
                if cell.column == 6 and cell.value == '差值':
                    for c in row:
                        if c.column >= 6:
                            c.font = red_font

        row_idx = 2
        while row_idx <= ws_summary.max_row:
            merge_end = row_idx + 6
            if merge_end <= ws_summary.max_row:
                for col in range(1, 6):
                    ws_summary.merge_cells(start_row=row_idx, start_column=col, end_row=merge_end, end_column=col)
            row_idx += 8

        output_filename = generate_temp_filename(f"周报_{week2_start.replace('-', '')}_{week2_end.replace('-', '')}")
        wb.save(output_filename)

        return output_filename

    finally:
        cursor.close()
        conn.close()


# ==================== 核心功能：生成月报 ====================
def generate_monthly_report(
    month1_start: str,
    month1_end: str,
    month2_start: str,
    month2_end: str,
    accounts: Optional[List[str]] = None
) -> str:
    """
    生成月报（两个月对比）
    复用周报逻辑
    """
    return generate_weekly_report(month1_start, month1_end, month2_start, month2_end, accounts)


# ==================== 核心功能：生成自定义报表 ====================
def generate_custom_report(
    period1_start: str,
    period1_end: str,
    period2_start: str,
    period2_end: str,
    accounts: Optional[List[str]] = None,
    shop_id: Optional[str] = None
) -> str:
    """
    生成自定义报表（两个自定义时间段对比，支持筛选门店）
    参数:
        accounts: 门店账号列表
        shop_id: 单个门店ID，用于筛选账号下的特定门店
    """
    return generate_weekly_report(period1_start, period1_end, period2_start, period2_end, accounts, shop_id)


# ==================== 核心功能：生成抖音商家日报 ====================
def generate_merchant_daily_report(data_date: str) -> str:
    """
    生成抖音商家每日数据报表
    参数:
        data_date: 数据日期，格式: 'YYYY-MM-DD'
    返回:
        生成的文件路径
    """
    db = get_db_pool()
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sql = """
        SELECT
            merchant_name, merchant_id, industry, category,
            cooperation_mode, follower_name,
            pay_amount, confirm_amount, refund_amount,
            item_pay_amount, room_pay_amount,
            ledger_commission, ledger_smc_commission,
            merchant_manage_score
        FROM douyin_merchant_daily
        WHERE data_date = %s
        ORDER BY merchant_id
        """
        cursor.execute(sql, [data_date])
        rows = cursor.fetchall()

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"商家数据_{data_date}"

        # 表头
        headers = [
            '商家名称', '商家ID', '行业', '类目',
            '合作模式', '跟进人',
            '支付GMV', '核销GMV', '退款GMV',
            '视频直接支付GMV', '直播支付GMV',
            '总预估佣金', '服务商预估佣金',
            '商家经营分'
        ]
        ws.append(headers)

        # 表头样式
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据行
        for row in rows:
            ws.append([
                row.get('merchant_name', ''),
                row.get('merchant_id', ''),
                row.get('industry', ''),
                row.get('category', ''),
                row.get('cooperation_mode', ''),
                row.get('follower_name', ''),
                round(float(row['pay_amount']), 2) if row.get('pay_amount') else 0,
                round(float(row['confirm_amount']), 2) if row.get('confirm_amount') else 0,
                round(float(row['refund_amount']), 2) if row.get('refund_amount') else 0,
                round(float(row['item_pay_amount']), 2) if row.get('item_pay_amount') else 0,
                round(float(row['room_pay_amount']), 2) if row.get('room_pay_amount') else 0,
                round(float(row['ledger_commission']), 2) if row.get('ledger_commission') else 0,
                round(float(row['ledger_smc_commission']), 2) if row.get('ledger_smc_commission') else 0,
                row.get('merchant_manage_score', ''),
            ])

        # 列宽设置
        col_widths = [20, 15, 12, 15, 12, 10, 14, 14, 14, 18, 16, 14, 16, 12]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 居中对齐 + 边框
        for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        apply_border(ws, 1, len(rows) + 1, 1, len(headers))

        output_filename = generate_temp_filename(f"抖音商家日报_{data_date.replace('-', '')}")
        wb.save(output_filename)

        return output_filename

    finally:
        cursor.close()
        conn.close()
